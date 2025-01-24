import os
import re
import asyncio
import statistics
import sys

from openpyxl import load_workbook
import pandas as pd

from config import (
    RESULTS_FILE,
    AUTOTESTS_FILE,
    EXPERIMENTS_DIR,
    TREND_PERMISSIBLE_ERROR,
    RELATIVE_ERROR,
    YEARS_TO_CHECK,
    TYPE_TO_EXECUTION_UUID,
)

QUANTITY_PREFIX = "[QUANTITY] "
QUALITY_PREFIX = "[QUALITY] "


def get_uuid_by_type(type: str) -> str:
    uuid = TYPE_TO_EXECUTION_UUID.get(type)
    if uuid is None:
        raise Exception(f"Нарушена связь TYPE to UUID для теста №{type}")
    return uuid


def get_file_path(files: list, execution_uuid: str, experiment_id: str):
    pattern = r"experiments\\results_(?P<execution_uuid>[a-zA-Z0-9-]+)_(?P<experiment_id>[a-zA-Z0-9-]+).xlsx"

    for file_path in files:
        match = re.match(pattern, file_path)
        if match:
            if (
                match.group("execution_uuid") == execution_uuid
                and match.group("experiment_id") == experiment_id
            ):
                return file_path

    return None


def prepare_trend_conditions(trend: str) -> list:
    result = []

    trend_conditions = [t.strip("() ").split(":") for t in trend.split(";")]

    for condition in trend_conditions:
        years_str, value = condition

        if "-" in years_str:
            start_year, end_year = map(int, years_str.split("-"))
            years = list(range(start_year, end_year + 1))
        else:
            years = [int(years_str)]

        value = int(value)

        for year in years:
            result.append((year, value))

    return result


def load_results_files() -> list:
    """Загружает все файлы результатов из папки экспериментов."""

    results_files = [
        os.path.join(EXPERIMENTS_DIR, file)
        for file in os.listdir(EXPERIMENTS_DIR)
        if file.startswith("results_") and file.endswith(".xlsx")
    ]

    if not results_files:
        raise FileNotFoundError(
            f"Не найдено ни одного файла с результатами в директории /{EXPERIMENTS_DIR}"
        )

    return results_files


def calculate_trend(base_value, compare_value) -> tuple:
    """Определяет тренд между двумя значениями (увеличение, уменьшение, неизменность)."""
    difference = compare_value - base_value

    trend = None
    if compare_value == base_value:
        trend = 0
    elif compare_value > base_value:
        trend = 1
    elif compare_value < base_value:
        trend = -1

    return difference, trend


def process_tests_common(
    tests_df, workbook, sheet_name, prefix, uuid_key, process_test
):
    """Общий процесс обработки тестов."""
    print(prefix, f"Началась обработка {prefix.lower()} тестов...")

    files_list = load_results_files()
    execution_uuid = get_uuid_by_type(uuid_key)

    base_file = get_file_path(files_list, execution_uuid, "0")
    base_df = pd.read_excel(base_file)

    results = []
    for index, test in tests_df.iterrows():
        experiment_id = str(test["id Теста"])

        print(prefix, f"PROCESS Experiment {experiment_id}")

        result = process_test(test, base_df, files_list, execution_uuid, experiment_id)

        sheet = workbook[sheet_name]

        # Update effect in quantitive tests
        for col_idx, value in enumerate(test):
            sheet.cell(row=index + 2, column=col_idx + 1, value=value)

        # Update result in Excel
        # sheet.cell(
        #     row=index + 2, column=tests_df.columns.get_loc("Итог") + 1, value=result
        # )

        results.append(
            {
                "execution_uuid": execution_uuid,
                "experiment_id": experiment_id,
                "result": result,
            }
        )

    return results


def process_qualitative_test(test, base_df, files_list, execution_uuid, experiment_id):
    """Обрабатывает один качественный тест."""
    result = True

    experiment_file = get_file_path(files_list, execution_uuid, experiment_id)
    if experiment_file is None:
        print(
            QUALITY_PREFIX,
            f"-- ERROR: Не найден файл с результатами эксперимента для №{experiment_id}",
        )
        result = False
    else:
        experiment_df = pd.read_excel(experiment_file)

        # Проверка "Взаимосвязь расчетов"
        linkage_test_result = True
        linkage = test["Взаимосвязь расчетов"]
        if linkage and pd.notna(linkage):
            # >|(id=14)|
            linked_sign = linkage[0]
            linked_id = linkage.split("id=")[1].split(")")[0]
            linked_file = get_file_path(files_list, execution_uuid, linked_id)
            linked_df = pd.read_excel(linked_file)

            linked_sum = linked_df["sum"].sum()
            current_sum = experiment_df["sum"].sum()

            linkage_test_result = False
            if linked_sign == ">":
                linkage_test_result = current_sum > linked_sum
            elif linked_sign == "<":
                linkage_test_result = current_sum < linked_sum
            else:
                print(
                    QUALITY_PREFIX,
                    f"-- ERROR: Символ '{linked_sign}' не предусмотрен для проверки взаимосвязи расчетов",
                )

            print(
                QUALITY_PREFIX,
                f"-- {'SUCCESS' if linkage_test_result else 'FAILED'}: linkage test {linkage}",
            )

        # Проверка "Тренд"
        base_df["year"] = pd.to_datetime(base_df["dt"]).dt.year
        experiment_df["year"] = pd.to_datetime(experiment_df["dt"]).dt.year

        trend_conditions = prepare_trend_conditions(test["Тренд"])
        trend_test_result = True
        for year, expected_trend in trend_conditions:
            base_value = base_df.loc[base_df["year"] == year, "sum"].values[0]
            compare_value = experiment_df.loc[
                experiment_df["year"] == year, "sum"
            ].values[0]

            if expected_trend != 0:
                difference, trend = calculate_trend(base_value, compare_value)
                if trend != expected_trend:
                    print(
                        QUALITY_PREFIX,
                        f"-- FAILED: trend test for year {year}, difference {difference} (expected trend {expected_trend})",
                    )
                    trend_test_result = False
                    break
            else:
                difference_percent = abs(base_value - compare_value) / base_value * 100
                if difference_percent > TREND_PERMISSIBLE_ERROR:
                    print(
                        QUALITY_PREFIX,
                        f"-- FAILED: trend test for year {year}, difference {difference_percent}%",
                    )
                    trend_test_result = False
                    break

            print(QUALITY_PREFIX, f"-- SUCCESS: trend test for year {year}")

        linkage_test_result = bool(linkage_test_result)
        trend_test_result = bool(trend_test_result)

        # Записываем результаты тестов
        test["Итог Взаимосвязь расчетов"] = linkage_test_result
        test["Итог Тренд"] = trend_test_result

        result = linkage_test_result and trend_test_result

    return result


def process_quantitative_test(test, base_df, files_list, execution_uuid, experiment_id):
    """Обрабатывает один количественный тест."""
    result = True

    experiment_file = get_file_path(files_list, execution_uuid, experiment_id)
    if experiment_file is None:
        print(
            QUANTITY_PREFIX,
            f"-- ERROR: Не найден файл с результатами эксперимента для №{experiment_id}",
        )
        result = False
    else:
        experiment_df = pd.read_excel(experiment_file)
        temp_df = experiment_df.copy()
        temp_df["difference"] = experiment_df["sum"] - base_df["sum"]
        temp_df["year"] = pd.to_datetime(temp_df["dt"]).dt.year

        errors = []  # Собираем все ошибки и считаем среднее количество

        for year in YEARS_TO_CHECK:
            effect_tnav = test[f"Эффект за {year} год по tNav"]
            effect_ml = temp_df.loc[temp_df["year"] == year, "difference"].values[0]

            relative_error = (
                100
                if effect_tnav == 0 and effect_ml != 0
                else ((effect_tnav - effect_ml) / effect_tnav) * 100
            )

            # Округление до 2 знака после запятой
            relative_error = round(relative_error, 2)

            # Заполняем таблицу с результатами
            test[f"Эффект за {year} год по ML"] = effect_ml
            test[f"Ошибка за {year} год"] = relative_error

            errors.append(relative_error)

        average_error = statistics.mean(errors)
        test["Средняя ошибка"] = average_error
        
        if abs(average_error) > RELATIVE_ERROR:
            print(
                QUANTITY_PREFIX,
                f"-- FAILED: average relative error {relative_error}% over the limit {RELATIVE_ERROR}%",
            )
            result = False
        else:
            print(QUANTITY_PREFIX, f"-- SUCCESS: for year {year}")

        test["Итог"] = bool(result)

    return result


def process_qualitative_tests(qualitative_df: pd.DataFrame, workbook) -> list:
    return process_tests_common(
        qualitative_df,
        workbook,
        "Список качественных автотестов",
        QUALITY_PREFIX,
        "quality",
        process_qualitative_test,
    )


def process_quantitative_tests(quantitative_df: pd.DataFrame, workbook) -> list:
    return process_tests_common(
        quantitative_df,
        workbook,
        "Список количественных автотесто",
        QUANTITY_PREFIX,
        "quantity",
        process_quantitative_test,
    )


def process_tests():
    """Обработка всех тестов и сохранение результатов в Excel."""
    workbook = load_workbook(AUTOTESTS_FILE)

    qualitative_df = pd.read_excel(
        AUTOTESTS_FILE, sheet_name="Список качественных автотестов"
    )
    quantitative_df = pd.read_excel(
        AUTOTESTS_FILE, sheet_name="Список количественных автотесто"
    )

    qualitative_results = process_qualitative_tests(qualitative_df, workbook)
    quantitative_results = process_quantitative_tests(quantitative_df, workbook)

    workbook.save(RESULTS_FILE)

    return qualitative_results, quantitative_results


from openpyxl import load_workbook


def process_statistics():
    # Открываем файл
    workbook = load_workbook(RESULTS_FILE)

    # Считываем данные из листов
    qualitative_sheet = workbook["Список качественных автотестов"]
    quantitative_sheet = workbook["Список количественных автотесто"]

    # Инициализируем переменные
    qualitative_results_trend = 0
    qualitative_results_relationship = 0
    qualitative_total_tests = 0

    qualitative_failed_trend = 0
    qualitative_failed_relationship = 0

    quantitative_total_tests = 0
    total_error = 0

    # Получаем названия столбцов на листе "Список качественных автотестов"
    qualitative_columns = {
        cell.value: col for col, cell in enumerate(qualitative_sheet[1], 1)
    }

    # Считываем данные о качественных тестах
    for row in qualitative_sheet.iter_rows(min_row=2, values_only=True):
        # Извлекаем значения по названию столбца
        trend_result = row[qualitative_columns.get("Итог Тренд") - 1]
        relationship_result = row[
            qualitative_columns.get("Итог Взаимосвязь расчетов") - 1
        ]

        qualitative_total_tests += 1

        if trend_result:
            qualitative_results_trend += 1
        else:
            qualitative_failed_trend += 1

        if relationship_result:
            qualitative_results_relationship += 1
        else:
            qualitative_failed_relationship += 1

    # Получаем названия столбцов на листе "Список количественных автотестов"
    quantitative_columns = {
        cell.value: col for col, cell in enumerate(quantitative_sheet[1], 1)
    }

    # Считываем данные о количественных тестах
    for row in quantitative_sheet.iter_rows(min_row=2, values_only=True):
        # Извлекаем значения по названию столбца
        error = row[quantitative_columns.get("Средняя ошибка") - 1]
        if error is not None:
            total_error += error
            quantitative_total_tests += 1

    # Рассчитываем проценты
    if qualitative_total_tests > 0:
        percent_completed_trend = (
            qualitative_results_trend / qualitative_total_tests
        ) * 100
        percent_completed_relationship = (
            qualitative_results_relationship / qualitative_total_tests
        ) * 100
        percent_failed_trend = (
            qualitative_failed_trend / qualitative_total_tests
        ) * 100
        percent_failed_relationship = (
            qualitative_failed_relationship / qualitative_total_tests
        ) * 100
    else:
        percent_completed_trend = 0
        percent_completed_relationship = 0
        percent_failed_trend = 0
        percent_failed_relationship = 0

    # Рассчитываем среднюю ошибку по количественным тестам
    if quantitative_total_tests > 0:
        quantitative_average_error = total_error / quantitative_total_tests
        print(
            f"Количество колич. тестов: {quantitative_total_tests}; Сумма всех ошибок: {total_error}; Средняя ошибка: {quantitative_average_error}"
        )
    else:
        quantitative_average_error = 0

    # Записываем статистику в лист "Статистика"

    stats_sheet = workbook["Статистика"]

    stats_sheet["B2"] = f"{percent_completed_trend:.2f}%"
    stats_sheet["B3"] = f"{percent_completed_relationship:.2f}%"
    stats_sheet["B4"] = f"{percent_failed_trend:.2f}%"
    stats_sheet["B5"] = f"{percent_failed_relationship:.2f}%"
    stats_sheet["B6"] = f"{quantitative_average_error:.2f}"

    workbook.save(RESULTS_FILE)
    print("Статистика успешно сохранена в файл.")


def run_stage_two():
    """
    Выполняет второй этап тестирования.
    """
    if not os.path.exists(EXPERIMENTS_DIR):
        raise FileNotFoundError(
            f"Директория {EXPERIMENTS_DIR} с результатами экспериментов не найдена."
        )

    process_tests()

    # Заполнение листа со статистикой
    process_statistics()

    print(f"Результаты тестирования сохранены в файле {RESULTS_FILE}")


if __name__ == "__main__":
    try:
        run_stage_two()
    except Exception as e:
        print(f"Ошибка выполнения второго этапа: {e}")
    finally:
        sys.exit(1)
